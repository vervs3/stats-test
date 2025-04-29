import requests
import pandas as pd
from datetime import datetime
import json
import argparse
import getpass
import os
import logging
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Basic logging setup
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger("jira_extraction")

# Try to import config
try:
    import config

    logger.info("Imported configuration successfully")
except ImportError:
    logger.warning("Failed to import config.py, will use command-line arguments instead")
    config = None

# Constants
CUTOFF_DATE = "2025-01-10T00:00:00.000+0000"
FIELDS_TO_FETCH = "summary,issuetype,created,timeoriginalestimate,subtasks,sprint"
HISTORY_EXPAND = "changelog"
ISSUE_TYPE_NEW_FEATURE = "New Feature"
TARGET_SPRINT_IDS = [14638, 14639, 14640, 14641]


def parse_args():
    parser = argparse.ArgumentParser(description="Extract Jira task estimations and compare current vs historical data")
    parser.add_argument("--jira-url", help="Jira instance URL, e.g., https://your-domain.atlassian.net")
    parser.add_argument("--filter-id", required=True, help="ID of the Jira filter to use")
    parser.add_argument("--token", help="Jira API token. If not provided, will use token from config.py")
    parser.add_argument("--output", default="jira_estimations.xlsx", help="Output Excel file name")
    parser.add_argument("--sprint-filter", action="store_true", help="Filter New Feature tasks by specific sprints")
    parser.add_argument("--all-tasks", action="store_true", help="Process all tasks, not just New Feature tasks")
    parser.add_argument("--debug", action="store_true", help="Enable debug logging")
    return parser.parse_args()


def get_jira_headers(token=None):
    api_token = token

    if not api_token and hasattr(config, 'api_token'):
        api_token = config.api_token
        logger.info("Using API token from config.py")

    if not api_token:
        logger.error("No API token provided. Use --token argument or define api_token in config.py")
        sys.exit(1)

    return {
        "Authorization": f"Bearer {api_token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }


def check_connection(jira_url, headers):
    try:
        logger.info("Checking Jira server availability...")
        try:
            session = requests.Session()
            resp = session.get(jira_url, timeout=10, allow_redirects=False)
            if resp.status_code >= 300 and resp.status_code < 400:
                logger.error(f"Server redirecting to: {resp.headers.get('Location', 'unknown')}")
                logger.error("VPN connection or NetScaler authentication may be required")
                return False
        except requests.RequestException as e:
            logger.error(f"Failed to connect to server: {e}")
            return False

        logger.info("Server available, checking token authentication...")
        response = requests.get(
            f"{jira_url}/rest/api/2/myself",
            headers=headers,
            timeout=10
        )
        logger.info(f"Response code: {response.status_code}")

        if response.status_code == 200:
            try:
                user_data = response.json()
                logger.info(f"Authentication successful! User: {user_data.get('displayName', 'unknown')}")
                return True
            except json.JSONDecodeError as e:
                logger.error(f"Error parsing JSON: {e}")
                return False
        else:
            logger.error(f"Authentication error. Code: {response.status_code}")
            return False
    except Exception as e:
        logger.error(f"Error checking connection: {e}")
        return False


def get_filter_jql(jira_url, filter_id, headers):
    filter_url = f"{jira_url}/rest/api/2/filter/{filter_id}"
    logger.info(f"Getting filter JQL from: {filter_url}")
    response = requests.get(filter_url, headers=headers)
    response.raise_for_status()
    jql = response.json().get('jql', '')
    logger.info(f"Retrieved JQL: {jql[:50]}...")
    return jql


def search_issues(jira_url, jql, headers):
    search_url = f"{jira_url}/rest/api/2/search"
    issues = []
    start_at = 0
    max_results = 50

    while True:
        params = {
            "jql": jql,
            "startAt": start_at,
            "maxResults": max_results,
            "fields": FIELDS_TO_FETCH,
            "expand": HISTORY_EXPAND
        }

        logger.info(f"Searching issues: startAt={start_at}, maxResults={max_results}")
        response = requests.get(search_url, headers=headers, params=params)
        response.raise_for_status()

        data = response.json()
        issues.extend(data["issues"])
        logger.info(f"Retrieved {len(data['issues'])} issues, total: {len(issues)}/{data['total']}")

        if start_at + max_results >= data["total"]:
            break

        start_at += max_results

    return issues


def get_subtasks(jira_url, issue_key, headers):
    issue_url = f"{jira_url}/rest/api/2/issue/{issue_key}"
    params = {
        "fields": f"subtasks,{FIELDS_TO_FETCH}",
        "expand": HISTORY_EXPAND
    }

    logger.info(f"Getting subtasks for issue: {issue_key}")
    response = requests.get(issue_url, headers=headers, params=params)
    response.raise_for_status()

    data = response.json()
    subtask_ids = [subtask["id"] for subtask in data["fields"].get("subtasks", [])]
    logger.info(f"Found {len(subtask_ids)} subtasks for {issue_key}")

    subtasks = []
    if subtask_ids:
        for subtask_id in subtask_ids:
            subtask_url = f"{jira_url}/rest/api/2/issue/{subtask_id}"
            logger.debug(f"Fetching subtask data for ID: {subtask_id}")
            subtask_response = requests.get(subtask_url, headers=headers, params={"expand": HISTORY_EXPAND})
            subtask_response.raise_for_status()
            subtasks.append(subtask_response.json())

    return subtasks


def convert_seconds_to_days(seconds):
    if seconds is None:
        return 0
    hours_per_day = 8
    return round(seconds / 3600 / hours_per_day, 2)


def get_original_estimate_at_date(issue, cutoff_date):
    # Get the original estimate of an issue at a specific date from history
    if not issue.get("changelog") or not issue["changelog"].get("histories"):
        created_date = issue["fields"]["created"]
        current_estimate = issue["fields"].get("timeoriginalestimate")

        if created_date <= cutoff_date:
            return current_estimate
        else:
            return 0

    if issue["fields"]["created"] > cutoff_date:
        return 0

    histories = sorted(issue["changelog"]["histories"], key=lambda x: x["created"])
    estimate = issue["fields"].get("timeoriginalestimate", 0)

    for history in reversed(histories):
        history_date = history["created"]

        if history_date <= cutoff_date:
            break

        for item in history["items"]:
            if item["field"] == "timeoriginalestimate":
                from_value = item.get("from")
                if from_value is not None and from_value != "":
                    estimate = int(from_value)

    return estimate


def get_sprint_info_at_date(issue, cutoff_date):
    # Extract sprint information from an issue that was valid at a specific date
    sprint_data = []
    sprint_ids_found = set()

    # Mapping of sprint names to IDs
    sprint_name_to_id = {
        "NBSS 25Q1": 14638,
        "NBSS 25Q2": 14639,
        "NBSS 25Q3": 14640,
        "NBSS 25Q4": 14641
    }

    current_sprints = get_current_sprint_info(issue)

    if issue["fields"]["created"] > cutoff_date:
        logger.debug(f"Issue {issue['key']} was created after cutoff date, no sprints to consider")
        return []

    if not issue.get("changelog") or not issue["changelog"].get("histories"):
        logger.debug(f"Issue {issue['key']} has no changelog, using current sprints")
        return current_sprints

    histories = sorted(issue["changelog"]["histories"], key=lambda x: x["created"])
    histories_before_cutoff = [h for h in histories if h["created"] <= cutoff_date]

    if not histories_before_cutoff and issue["fields"]["created"] <= cutoff_date:
        logger.debug(f"Issue {issue['key']} has no changelog before cutoff but was created before")
        return current_sprints

    for history in histories_before_cutoff:
        for item in history.get("items", []):
            if item.get("field") == "Sprint":
                sprint_str = item.get("toString", "")
                logger.debug(f"Found sprint change in history: '{sprint_str}'")

                if sprint_str:
                    sprint_data = []
                    sprint_ids_found = set()

                    # Find sprint IDs in format [ XXXX ]
                    sprint_ids = re.findall(r'\[\s*(\d+)\s*\]', sprint_str)

                    # Find sprint names in format "NBSS XXqX"
                    sprint_names = re.findall(r'(NBSS\s+\d+Q[1-4])', sprint_str)
                    if not sprint_names:
                        sprint_names = [name.strip() for name in re.split(r'[,\s]+', sprint_str) if name.strip()]

                    logger.debug(f"Extracted sprint IDs: {sprint_ids}, names: {sprint_names}")

                    # Process combined format "NAME [ ID ]"
                    sprint_blocks = re.findall(r'([^\[\]]+)\[\s*(\d+)\s*\]', sprint_str)

                    if sprint_blocks:
                        for name, sprint_id in sprint_blocks:
                            try:
                                sprint_id_int = int(sprint_id)
                                if sprint_id_int not in sprint_ids_found:
                                    sprint_ids_found.add(sprint_id_int)
                                    sprint_data.append({
                                        "id": sprint_id_int,
                                        "name": name.strip(),
                                        "state": "unknown"
                                    })
                            except ValueError:
                                logger.warning(f"Error parsing sprint ID '{sprint_id}'")
                    elif sprint_ids:
                        for sprint_id in sprint_ids:
                            try:
                                sprint_id_int = int(sprint_id)
                                if sprint_id_int not in sprint_ids_found:
                                    sprint_ids_found.add(sprint_id_int)
                                    sprint_data.append({
                                        "id": sprint_id_int,
                                        "name": f"Sprint {sprint_id}",
                                        "state": "unknown"
                                    })
                            except ValueError:
                                logger.warning(f"Error parsing sprint ID '{sprint_id}'")
                    elif sprint_names:
                        for name in sprint_names:
                            name_clean = name.strip()
                            if name_clean in sprint_name_to_id:
                                sprint_id_int = sprint_name_to_id[name_clean]
                                if sprint_id_int not in sprint_ids_found:
                                    sprint_ids_found.add(sprint_id_int)
                                    sprint_data.append({
                                        "id": sprint_id_int,
                                        "name": name_clean,
                                        "state": "unknown"
                                    })
                            else:
                                quarter_match = re.search(r'NBSS\s+(\d+)Q([1-4])', name_clean)
                                if quarter_match:
                                    year = quarter_match.group(1)
                                    quarter = quarter_match.group(2)

                                    if f"NBSS {year}Q{quarter}" in sprint_name_to_id:
                                        sprint_id_int = sprint_name_to_id[f"NBSS {year}Q{quarter}"]

                                        if sprint_id_int not in sprint_ids_found:
                                            sprint_ids_found.add(sprint_id_int)
                                            sprint_data.append({
                                                "id": sprint_id_int,
                                                "name": name_clean,
                                                "state": "unknown"
                                            })

    if not sprint_data:
        logger.debug(f"No sprint changes found in history for {issue['key']}, using current sprints")
        sprint_data = current_sprints

    return sprint_data


def get_current_sprint_info(issue):
    sprint_data = []
    sprint_ids_found = set()

    if "sprint" in issue["fields"]:
        sprints = issue["fields"]["sprint"]
        if sprints:
            if isinstance(sprints, list):
                for sprint in sprints:
                    sprint_id = sprint.get("id")
                    if sprint_id and sprint_id not in sprint_ids_found:
                        sprint_ids_found.add(sprint_id)
                        sprint_data.append({
                            "id": sprint_id,
                            "name": sprint.get("name", f"Sprint {sprint_id}"),
                            "state": sprint.get("state", "unknown"),
                        })
            else:
                sprint_id = sprints.get("id")
                if sprint_id and sprint_id not in sprint_ids_found:
                    sprint_ids_found.add(sprint_id)
                    sprint_data.append({
                        "id": sprint_id,
                        "name": sprints.get("name", f"Sprint {sprint_id}"),
                        "state": sprints.get("state", "unknown"),
                    })

    for field_name, field_value in issue["fields"].items():
        if field_name.startswith("customfield_") and field_value is not None:
            if isinstance(field_value, list):
                for item in field_value:
                    if isinstance(item, str) and "sprint" in item.lower():
                        try:
                            sprint_id_match = re.search(r'id=(\d+)', item)
                            if sprint_id_match:
                                sprint_id = int(sprint_id_match.group(1))
                                if sprint_id not in sprint_ids_found:
                                    sprint_ids_found.add(sprint_id)

                                    sprint_name_match = re.search(r'name=([^,]+)', item)
                                    sprint_state_match = re.search(r'state=([^,]+)', item)

                                    sprint_data.append({
                                        "id": sprint_id,
                                        "name": sprint_name_match.group(
                                            1) if sprint_name_match else f"Sprint {sprint_id}",
                                        "state": sprint_state_match.group(1) if sprint_state_match else "unknown",
                                    })
                        except Exception as e:
                            logger.warning(f"Error parsing sprint string '{item}': {e}")

    return sprint_data


def process_issues(issues, jira_url, headers, cutoff_date, sprint_filter=False, all_tasks=False):
    # Process all issues and extract required data
    results = []
    total_processed = 0
    total_included = 0

    for issue in issues:
        issue_key = issue["key"]
        issue_type = issue["fields"]["issuetype"]["name"]
        total_processed += 1

        if not all_tasks and issue_type != ISSUE_TYPE_NEW_FEATURE:
            logger.info(f"Skipping {issue_key} as it is not a New Feature task")
            continue

        if sprint_filter:
            sprints = get_sprint_info_at_date(issue, cutoff_date)
            sprint_ids = [sprint.get("id") for sprint in sprints]

            logger.info(f"Found sprint IDs for {issue_key} at cutoff date: {sprint_ids}")

            if not any(sprint_id in TARGET_SPRINT_IDS for sprint_id in sprint_ids):
                logger.info(f"Skipping {issue_key} as it does not belong to target sprints.")
                continue
            else:
                logger.info(f"Including {issue_key} as it belongs to target sprints: {sprint_ids}")
        else:
            sprints = get_current_sprint_info(issue)

        total_included += 1
        logger.info(f"Processing {issue_key}: {issue['fields']['summary']} (Type: {issue_type})")

        issue_current_estimate = issue["fields"].get("timeoriginalestimate", 0) or 0
        issue_historical_estimate = get_original_estimate_at_date(issue, cutoff_date) or 0

        subtasks = get_subtasks(jira_url, issue_key, headers)

        current_subtask_estimates = 0
        historical_subtask_estimates = 0
        subtask_data = []

        for subtask in subtasks:
            subtask_key = subtask["key"]
            subtask_created = subtask["fields"]["created"]
            subtask_current_estimate = subtask["fields"].get("timeoriginalestimate", 0) or 0

            subtask_historical_estimate = get_original_estimate_at_date(subtask, cutoff_date)
            if subtask_historical_estimate is None:
                subtask_historical_estimate = 0

            if subtask_created <= cutoff_date:
                historical_subtask_estimates += subtask_historical_estimate

            current_subtask_estimates += subtask_current_estimate

            if sprint_filter:
                subtask_sprints = get_sprint_info_at_date(subtask, cutoff_date)
            else:
                subtask_sprints = get_current_sprint_info(subtask)

            subtask_sprint_names = [sprint.get("name") for sprint in subtask_sprints]

            subtask_data.append({
                "issue_key": subtask_key,
                "summary": subtask["fields"]["summary"],
                "issue_type": subtask["fields"]["issuetype"]["name"],
                "created": subtask_created,
                "historical_estimate_seconds": subtask_historical_estimate,
                "historical_estimate_days": convert_seconds_to_days(subtask_historical_estimate),
                "current_estimate_seconds": subtask_current_estimate,
                "current_estimate_days": convert_seconds_to_days(subtask_current_estimate),
                "parent_key": issue_key,
                "level": 1,  # Subtasks are level 1, parents are level 0
                "sprints": subtask_sprint_names
            })

        if issue_type == ISSUE_TYPE_NEW_FEATURE:
            final_current_estimate = current_subtask_estimates
            final_historical_estimate = historical_subtask_estimates
        else:
            final_current_estimate = issue_current_estimate
            final_historical_estimate = issue_historical_estimate

        issue_sprint_names = [sprint.get("name") for sprint in sprints]

        results.append({
            "issue_key": issue_key,
            "summary": issue["fields"]["summary"],
            "issue_type": issue_type,
            "created": issue["fields"]["created"],
            "historical_estimate_seconds": final_historical_estimate,
            "historical_estimate_days": convert_seconds_to_days(final_historical_estimate),
            "current_estimate_seconds": final_current_estimate,
            "current_estimate_days": convert_seconds_to_days(final_current_estimate),
            "parent_key": None,
            "level": 0,  # Parent issues are level 0
            "sprints": issue_sprint_names
        })

        results.extend(subtask_data)

    logger.info(f"Processed {total_processed} issues, included {total_included} in results")
    return results


def create_excel_report(data, output_file, sprint_filter=False):
    # Create an Excel report with the extracted data
    wb = Workbook()
    ws = wb.active
    ws.title = "Jira Estimations"

    headers = [
        "Issue Key",
        "Summary",
        "Issue Type",
        "Created Date",
        "Оценка до 10 января 2025 (дни)",
        "Оценка в настоящий момент (дни)",
        "Спринты"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    row_num = 2
    totals = {}

    for item in data:
        prefix = "    " * item["level"] if item["level"] > 0 else ""

        ws.cell(row=row_num, column=1).value = item["issue_key"]
        ws.cell(row=row_num, column=2).value = f"{prefix}{item['summary']}"
        ws.cell(row=row_num, column=3).value = item["issue_type"]
        ws.cell(row=row_num, column=4).value = item["created"][:10]  # Just the date part
        ws.cell(row=row_num, column=5).value = item["historical_estimate_days"]
        ws.cell(row=row_num, column=6).value = item["current_estimate_days"]
        ws.cell(row=row_num, column=7).value = ", ".join(item.get("sprints", []))

        if item["level"] == 0:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).font = Font(bold=True)

            issue_type = item["issue_type"]
            if issue_type not in totals:
                totals[issue_type] = {
                    "count": 0,
                    "historical": 0,
                    "current": 0
                }

            totals[issue_type]["count"] += 1
            totals[issue_type]["historical"] += item["historical_estimate_days"]
            totals[issue_type]["current"] += item["current_estimate_days"]

        row_num += 1

    row_num += 1

    for issue_type, stats in totals.items():
        title_prefix = "Итого по "
        if sprint_filter and issue_type == ISSUE_TYPE_NEW_FEATURE:
            title_prefix = "Итого по New Feature в целевых спринтах: "

        ws.cell(row=row_num, column=1).value = f"{title_prefix}{issue_type}:"
        ws.cell(row=row_num, column=2).value = f"{stats['count']} задач"
        ws.cell(row=row_num, column=5).value = stats["historical"]
        ws.cell(row=row_num, column=6).value = stats["current"]
        ws.cell(row=row_num, column=7).value = f"Изменение: {stats['current'] - stats['historical']:.2f} дней"

        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).font = Font(bold=True)
            ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE",
                                                                    fill_type="solid")

        row_num += 1

    total_issues = sum(stats["count"] for stats in totals.values())
    total_historical = sum(stats["historical"] for stats in totals.values())
    total_current = sum(stats["current"] for stats in totals.values())

    row_num += 1
    title = "ОБЩИЙ ИТОГ:"
    if sprint_filter:
        title = "ОБЩИЙ ИТОГ ПО ЗАДАЧАМ В ЦЕЛЕВЫХ СПРИНТАХ:"

    ws.cell(row=row_num, column=1).value = title
    ws.cell(row=row_num, column=2).value = f"Все {total_issues} задачи"
    ws.cell(row=row_num, column=5).value = total_historical
    ws.cell(row=row_num, column=6).value = total_current
    ws.cell(row=row_num, column=7).value = f"Изменение: {total_current - total_historical:.2f} дней"

    for col_num in range(1, len(headers) + 1):
        ws.cell(row=row_num, column=col_num).font = Font(bold=True)
        ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD",
                                                                fill_type="solid")

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
    logger.info(f"Report saved as {output_file}")


def main():
    args = parse_args()

    if args.debug:
        logger.setLevel(logging.DEBUG)
        logger.debug("Debug logging enabled")

    jira_url = args.jira_url
    if not jira_url and hasattr(config, 'jira_url'):
        jira_url = config.jira_url
        logger.info(f"Using Jira URL from config.py: {jira_url}")

    if not jira_url:
        jira_url = "https://jira.nexign.com"
        logger.info(f"Using default Jira URL: {jira_url}")

    headers = get_jira_headers(args.token)

    try:
        if not check_connection(jira_url, headers):
            logger.warning("Connection check failed, but will try to continue.")

        jql = get_filter_jql(jira_url, args.filter_id, headers)

        logger.info(f"Searching for issues using filter ID {args.filter_id}...")
        issues = search_issues(jira_url, jql, headers)
        logger.info(f"Found {len(issues)} issues")

        logger.info(f"Processing issues with sprint filter: {args.sprint_filter}, all tasks: {args.all_tasks}")
        results = process_issues(issues, jira_url, headers, CUTOFF_DATE,
                                 sprint_filter=args.sprint_filter,
                                 all_tasks=args.all_tasks)

        logger.info(f"Processed {len(results)} issues (including subtasks)")

        create_excel_report(results, args.output, sprint_filter=args.sprint_filter)

    except requests.exceptions.HTTPError as e:
        logger.error(f"HTTP Error: {e}")
        if e.response.status_code == 401:
            logger.error("Authentication failed. Please check your API token.")
        elif e.response.status_code == 404:
            logger.error("Filter not found. Please check the filter ID.")
    except Exception as e:
        logger.error(f"Error: {e}", exc_info=True)


if __name__ == "__main__":
    main()